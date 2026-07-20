import ipaddress
import re
import threading
import uuid
from io import BytesIO
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.db.models.functions import TruncDay, TruncHour
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from openpyxl import load_workbook

from core.autenticacion import grupo_requerido
from core.models import Cliente, Logs, Plan
from .forms import ClienteForm, FiltroClientes


# Este metodo se encarga de mostrar la lista de clientes registrados, 
# con la posibilidad de aplicar filtros por nombre del cliente y cedula.
@login_required
@grupo_requerido('asistente_administrativo')
def gestion_cliente(request):
    todos_clientes = Cliente.objects.filter(borrado=False).order_by('nombre')

    if request.method == 'POST':
        filtro = FiltroClientes(request.POST)
        if filtro.is_valid():
            nombreCliente = filtro.cleaned_data.get('nombreCliente')
            estado_seleccionado = filtro.cleaned_data.get('estado')

            # Filtro por texto (Nombre O Cédula)
            if nombreCliente:
                nombreCliente = nombreCliente.strip()
                todos_clientes = todos_clientes.filter(
                    Q(nombre__icontains=nombreCliente) | 
                    Q(cedula__icontains=nombreCliente)
                )

            if estado_seleccionado:
                todos_clientes = todos_clientes.filter(estado=estado_seleccionado)
    else:
        filtro = FiltroClientes()

    paginator = Paginator(todos_clientes, 10)
    
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    
    page_number = request.GET.get('page')
    clientes = paginator.get_page(page_number)
        
    return render(request, 'gestion_clientes.html', {
        'clientes': clientes, 
        'filtros': filtro, 
        'query_string': query_params.urlencode() 
    })


# Este metodo se encarga de registrar un nuevo cliente, 
#si las entradas son validas, actualizamos el saldo del cliente, su estado (pendiente o exonerado) y la fecha. 
# Registra un log detallado del nuevo cliente registrado.
@login_required
@grupo_requerido('asistente_administrativo')
def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES)
        if form.is_valid():
            
            cliente = form.save(commit=False)
            cliente.fechaRegistro = timezone.now()

            if form.cleaned_data.get('exonerar_cliente'):
                cliente.estado = 'Exonerado'
            else:
                cliente.estado = 'Solvente'         
                cliente.borrado = False       
            
            cliente.save()   

            Logs.objects.create(
                idPersonal=request.user,
                mensaje=f"""
                Registró al nuevo cliente {cliente.nombre} (Cédula: {cliente.cedula}).
                Celular: {cliente.celular}
                Email: {cliente.email}
                Direccion: {cliente.direccion}
                Plan: {cliente.idPlan.plan} 
                Direccion Ip: {cliente.direccionIP}
                Estado: {cliente.estado}
                """,
                modulo="Gestión de Clientes",
                error=False,
                fecha=timezone.now()
            )

            return redirect('gestion_clientes') 
    else:
        form = ClienteForm() 

    return render(request, 'crear_cliente.html', {'form': form})

# Este metodo se encarga de modificar un cliente existente, 
# si las entradas son validas, actualizamos el saldo del cliente, su estado (pendiente o exonerado) y la fecha. 
# Registra un log detallado del nuevo cliente registrado.
@login_required
@grupo_requerido('asistente_administrativo')
def modificar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        cliente_viejo = Cliente.objects.get(id=id)
        form = ClienteForm(request.POST, instance=cliente)
        
        if form.is_valid():

            cliente = form.save(commit=False)

            if form.cleaned_data.get('exonerar_cliente'):
                cliente.estado = 'Exonerado'
                cliente.saldo = 0.00
            else: 
                if cliente.estado == 'Exonerado':
                    cliente.estado = 'Pendiente'
                    cliente.saldo = cliente.idPlan.precioUSD

            cliente = form.save()
            lista_cambios = []

            for campo, valor_nuevo in cliente.__dict__.items():

                if campo.startswith('_') or campo in ['id', 'borrado']:
                    continue
                
                valor_viejo = getattr(cliente_viejo, campo)
                
                if valor_nuevo != valor_viejo:
                    lista_cambios.append(f"{campo}: {valor_viejo} => {valor_nuevo}")

            if lista_cambios:
                mensaje_log = f"Modificó al cliente {cliente.nombre} (ID: {cliente.id}). Cambios realizados:\n\n" + "\n".join(lista_cambios)
            else:
                mensaje_log = f"El operador guardó al cliente {cliente.nombre} (ID: {cliente.id}) sin realizar cambios."

            Logs.objects.create(
                idPersonal=request.user,
                mensaje=mensaje_log,
                modulo="Gestión de Clientes",
                error=False,
                fecha=timezone.now()
            )

            return redirect('gestion_clientes')
            
    else:
        form = ClienteForm(instance=cliente)
        
    return render(request, 'modificar_cliente.html', {'form': form, 'cliente': cliente})

# Este metodo se encarga de borrar un cliente existente, 
# si las entradas son validas, actualizamos su estado de borrado logico y lo cambiamos a TRUE, esto pondra al cliente como eliminado. 
# Registra un log detallado del nuevo cliente registrado.
@login_required
@grupo_requerido('asistente_administrativo')
def borrar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        
        cliente.borrado = True

        cliente.save()

        Logs.objects.create(
                idPersonal=request.user,
                mensaje=f"""
                Se Elimino al cliente {cliente.nombre} (Cédula: {cliente.cedula}).
                Celular: {cliente.celular}
                Email: {cliente.email}
                Direccion: {cliente.direccion}
                Plan: {cliente.idPlan.plan} 
                Direccion Ip:{cliente.direccionIP}
                Estado: {cliente.estado}
                """,
                modulo="Gestión de Clientes",
                error=False,
                fecha=timezone.now()
            )
        
        return redirect('gestion_clientes')
    return render(request, 'confirmar_borrar.html', {'cliente': cliente})

# Este método devuelve los totales de clientes por estado para alimentar los indicadores del dashboard.
def api_tarjetas_dashboard(request):
    # Conteo global absoluto de la base de datos sin importar filtros temporales
    solventes = Cliente.objects.filter(estado='Solvente', borrado=False).count()
    exonerados = Cliente.objects.filter(estado='Exonerado', borrado=False).count()
    pendientes = Cliente.objects.filter(estado='Pendiente', borrado=False).count()
    desconectados = Cliente.objects.filter(estado='Desconectado', borrado=False).count()
    
    data = {
        'solventes': solventes,
        'exonerados': exonerados,
        'pendientes': pendientes,
        'desconectados': desconectados,
    }
    
    return JsonResponse(data)

# Este método prepara los datos para los gráficos del dashboard con base en logs y estado de cobranza.
def api_graficos_dashboard(request):
    filtro = request.GET.get('filtro', 'actualmente')
    ahora = timezone.now()
    
    # --- 1. FILTRADO PARA EL HISTÓRICO DE LOGS ---
    logs_base = Logs.objects.all()
    
    if filtro == '7dias':
        fecha_inicio_logs = ahora - timedelta(days=7)
        logs_base = logs_base.filter(fecha__gte=fecha_inicio_logs)
        logs_query = logs_base.annotate(periodo=TruncDay('fecha'))
    elif filtro == 'mes':
        inicio_mes_logs = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        logs_base = logs_base.filter(fecha__gte=inicio_mes_logs)
        logs_query = logs_base.annotate(periodo=TruncDay('fecha'))
    else:
        # 'actualmente' -> Últimas 24 horas reales agrupadas por HORA
        fecha_inicio_logs = ahora - timedelta(hours=24)
        logs_base = logs_base.filter(fecha__gte=fecha_inicio_logs)
        logs_query = logs_base.annotate(periodo=TruncHour('fecha'))

    # Agrupación y formateo de los logs procesados según el periodo anotado
    logs_query = (
        logs_query
        .values('periodo')
        .annotate(
            exitos=Count('id', filter=Q(error=False)),
            errores=Count('id', filter=Q(error=True))
        )
        .order_by('periodo')
    )
    
    labels_logs = []
    series_exitos = []
    series_errores = []
    
    dias_es = {
        'Mon': 'Lun', 'Tue': 'Mar', 'Wed': 'Mie', 
        'Thu': 'Jue', 'Fri': 'Vie', 'Sat': 'Sab', 'Sun': 'Dom'
    }
    
    for log in logs_query:
        if log['periodo']:
            if filtro == 'actualmente':
                # Si es por horas, formateamos como "14:00"
                label_final = log['periodo'].strftime('%H:%M')
            elif filtro == 'mes':
                # Si es el mes entero, formato "DD/MM"
                label_final = log['periodo'].strftime('%d/%m')
            else:
                # Si son 7 días, el nombre del día traducido
                dia_en = log['periodo'].strftime('%a')
                label_final = dias_es.get(dia_en, dia_en)
                
            labels_logs.append(label_final)
            series_exitos.append(log['exitos'])
            series_errores.append(log['errores'])

    # --- 2. FILTRADO PARA EL ESTADO DE COBRANZAS (DONA) ---
    # Refleja el estado financiero global en tiempo real
    clientes_cobranzas = Cliente.objects.filter(borrado=False, estado__in=['Solvente', 'Pendiente'])

    cobranzas_query = (
        clientes_cobranzas
        .values('estado')
        .annotate(total=Count('id'))
    )
    
    distribucion_cobranzas = {item['estado']: item['total'] for item in cobranzas_query}
    
    if 'Solvente' not in distribucion_cobranzas: distribucion_cobranzas['Solvente'] = 0
    if 'Pendiente' not in distribucion_cobranzas: distribucion_cobranzas['Pendiente'] = 0

    json_final = {
        'historico_logs': {
            'labels': labels_logs,
            'exitos': series_exitos,
            'errores': series_errores
        },
        'estado_cobranzas': distribucion_cobranzas
    }

    return JsonResponse(json_final, json_dumps_params={'ensure_ascii': False})

# Este metodo se encarga de mostrar la informacion detallada de un cliente especifico,
# cargando tanto sus datos personales como los del servicio contratado de forma de solo lectura.
@login_required
@grupo_requerido('asistente_administrativo')
def detalles_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    return render(request, 'detalles_cliente.html', {'cliente': cliente})


# Este método inicia la carga masiva de clientes desde un archivo Excel y deja un identificador para consultar el progreso.
@login_required
@grupo_requerido('asistente_administrativo')
def carga_clientes(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            return render(request, 'carga_clientes.html', {'error': 'Selecciona un archivo .xlsx para continuar.'})

        request_id = uuid.uuid4().hex
        archivo_bytes = archivo.read()

        hilo = threading.Thread(
            target=_procesar_importacion_en_segundo_plano,
            args=(request_id, archivo_bytes, request.user),
            daemon=True,
        )
        hilo.start()

        return render(request, 'carga_clientes.html', {
            'request_id': request_id,
            'archivo_nombre': archivo.name,
            'estado_inicial': 'Iniciando la carga...',
        })

    return render(request, 'carga_clientes.html')


# Este método devuelve el estado actual de la importación para que la interfaz pueda actualizar la barra de carga.
@login_required
@grupo_requerido('asistente_administrativo')
def progreso_importacion(request, request_id):
    estado = _estado_importacion(request_id)
    if not estado:
        return JsonResponse({'estado': 'esperando', 'paso': 'Esperando inicio...', 'progreso': 0, 'mensaje': 'La importación aún no ha comenzado.'})

    return JsonResponse(estado)

# Esta función normaliza texto proveniente del archivo Excel, eliminando espacios innecesarios.
def _normalizar_texto(valor):
    if valor is None:
        return ''
    return str(valor).strip()


# Esta función convierte los encabezados del Excel a un formato uniforme para facilitar su lectura.
def _normalizar_header(valor):
    return re.sub(r'[^a-z0-9]+', '', _normalizar_texto(valor).lower())


# Esta función limpia la cédula importada para dejarla en un formato numérico consistente.
def _limpiar_cedula(valor):
    texto = _normalizar_texto(valor).replace('-', '').replace('.', '').replace(' ', '')
    return texto


# Esta función limpia el número de celular importado para dejarlo listo para su almacenamiento.
def _limpiar_celular(valor):
    texto = _normalizar_texto(valor).replace('-', '').replace('.', '').replace(' ', '').replace('(', '').replace(')', '')
    return texto


# Este método procesa un archivo Excel y registra o actualiza los clientes en la base de datos, además de publicar su avance.
def importar_clientes_desde_excel(archivo, usuario, request_id=None):
    contenido = archivo.read() if hasattr(archivo, 'read') else archivo
    workbook = load_workbook(filename=BytesIO(contenido), data_only=True)
    sheet = workbook.active

    filas = list(sheet.iter_rows(values_only=True))
    if not filas:
        _guardar_estado_importacion(request_id, {
            'estado': 'completado',
            'paso': 'Carga finalizada',
            'progreso': 100,
            'total': 0,
            'procesados': 0,
            'creados': 0,
            'actualizados': 0,
            'errores': 0,
            'mensaje': 'El archivo está vacío.',
        })
        return {'creados': 0, 'actualizados': 0, 'errores': 0, 'total': 0, 'mensaje': 'El archivo está vacío.'}

    headers = [_normalizar_header(celda) for celda in filas[0]]
    total = 0
    creados = 0
    actualizados = 0
    errores = 0

    for fila in filas[1:]:
        if not any(_normalizar_texto(celda) not in ['', None] for celda in fila):
            continue
        total += 1

    _guardar_estado_importacion(request_id, {
        'estado': 'procesando',
        'paso': 'Leyendo archivo...',
        'progreso': 10,
        'total': total,
        'procesados': 0,
        'creados': 0,
        'actualizados': 0,
        'errores': 0,
        'mensaje': 'Se preparan las filas del archivo para la importación.',
    })

    procesados = 0

    for indice, fila in enumerate(filas[1:], start=2):
        if not any(_normalizar_texto(celda) not in ['', None] for celda in fila):
            continue

        procesados += 1

        datos = {}
        for posicion, valor in enumerate(headers):
            datos[valor] = fila[posicion] if posicion < len(fila) else ''

        nombre = _normalizar_texto(datos.get('nombre') or datos.get('clientenombre'))
        cedula = _limpiar_cedula(datos.get('cedula') or datos.get('rif'))
        celular = _limpiar_celular(datos.get('celular') or datos.get('telefono'))
        direccion = _normalizar_texto(datos.get('direccion') or datos.get('direccioncliente'))
        email = _normalizar_texto(datos.get('email') or datos.get('correo'))
        direccion_ip = _normalizar_texto(datos.get('direccionip') or datos.get('ip'))
        plan_nombre = _normalizar_texto(datos.get('plan') or datos.get('nombreplan'))
        estado = _normalizar_texto(datos.get('estado'))
        saldo_texto = _normalizar_texto(datos.get('saldo'))

        if not nombre or not cedula:
            errores += 1
            continue

        if len(cedula) < 6:
            cedula = cedula.zfill(6)
        elif len(cedula) > 9:
            cedula = cedula[:9]

        if celular:
            if len(celular) < 11:
                celular = celular.zfill(11)
            elif len(celular) > 11:
                celular = celular[:11]

        if email and '@' not in email:
            errores += 1
            continue

        try:
            ipaddress.ip_address(direccion_ip)
        except ValueError:
            if direccion_ip:
                errores += 1
                continue

        plan = None
        if plan_nombre:
            plan = Plan.objects.filter(plan__iexact=plan_nombre, borrado=False).first()

        if not plan:
            plan = Plan.objects.filter(pk=1, borrado=False).first()

        if not plan:
            errores += 1
            continue

        cliente_existente = Cliente.objects.filter(cedula=cedula, borrado=False).first()
        if cliente_existente:
            cliente = cliente_existente
            actualizados += 1
        else:
            cliente = Cliente()
            cliente.cedula = cedula
            creados += 1

        cliente.idPlan = plan
        cliente.nombre = nombre.upper()
        cliente.celular = celular
        cliente.direccion = direccion.upper()
        cliente.email = email.lower()
        cliente.direccionIP = direccion_ip or cliente.direccionIP

        if estado:
            cliente.estado = estado
        else:
            cliente.estado = 'Pendiente'

        try:
            cliente.saldo = float(saldo_texto) if saldo_texto else plan.precioUSD
        except (TypeError, ValueError):
            cliente.saldo = plan.precioUSD

        cliente.borrado = False
        cliente.fechaRegistro = cliente.fechaRegistro or timezone.now()
        cliente.save()

        if request_id is not None:
            progreso = 10 if total <= 0 else int(10 + (procesados / total) * 85)
            _guardar_estado_importacion(request_id, {
                'estado': 'procesando',
                'paso': f'Procesando fila {procesados} de {total}',
                'progreso': min(progreso, 95),
                'total': total,
                'procesados': procesados,
                'creados': creados,
                'actualizados': actualizados,
                'errores': errores,
                'mensaje': f'Fila {procesados} de {total} procesada.',
            })

    Logs.objects.create(
        idPersonal=usuario,
        modulo='Gestión de Clientes',
        mensaje=(
            f'Importación masiva de clientes finalizada. '
            f'Total procesadas: {total}. Creados: {creados}. Actualizados: {actualizados}. '
            f'Errores: {errores}.'
        ),
        error=errores > 0,
        fecha=timezone.now(),
    )

    _guardar_estado_importacion(request_id, {
        'estado': 'completado',
        'paso': 'Carga finalizada',
        'progreso': 100,
        'total': total,
        'procesados': total,
        'creados': creados,
        'actualizados': actualizados,
        'errores': errores,
        'mensaje': 'Importación completada.',
    })

    return {
        'creados': creados,
        'actualizados': actualizados,
        'errores': errores,
        'total': total,
        'mensaje': 'Importación completada.',
    }


# Esta función obtiene el estado actual de una importación masiva en curso.
def _estado_importacion(request_id):
    return cache.get(f'importacion_clientes:{request_id}')


# Esta función guarda el estado de progreso de la importación para que la vista pueda consultarlo.
def _guardar_estado_importacion(request_id, estado):
    if request_id is None:
        return
    cache.set(f'importacion_clientes:{request_id}', estado, timeout=100)

# Este método ejecuta la carga masiva en segundo plano y actualiza el progreso en memoria.
def _procesar_importacion_en_segundo_plano(request_id, archivo_bytes, usuario):
    _guardar_estado_importacion(request_id, {
        'estado': 'procesando',
        'paso': 'Iniciando importación...',
        'progreso': 5,
        'total': 0,
        'procesados': 0,
        'creados': 0,
        'actualizados': 0,
        'errores': 0,
        'mensaje': 'Preparando el archivo para la carga.',
    })

    try:
        import_resultado = importar_clientes_desde_excel(BytesIO(archivo_bytes), usuario, request_id=request_id)
        _guardar_estado_importacion(request_id, {
            'estado': 'completado',
            'paso': 'Carga finalizada',
            'progreso': 100,
            'total': import_resultado.get('total', 0),
            'procesados': import_resultado.get('total', 0),
            'creados': import_resultado.get('creados', 0),
            'actualizados': import_resultado.get('actualizados', 0),
            'errores': import_resultado.get('errores', 0),
            'mensaje': import_resultado.get('mensaje', 'Importación completada.'),
        })
    except Exception as exc:
        _guardar_estado_importacion(request_id, {
            'estado': 'error',
            'paso': 'Error de importación',
            'progreso': 100,
            'total': 0,
            'procesados': 0,
            'creados': 0,
            'actualizados': 0,
            'errores': 1,
            'mensaje': f'No fue posible completar la importación: {exc}',
        })
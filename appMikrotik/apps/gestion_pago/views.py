from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
from core.models import Pago, Logs, Cliente
from .forms import PagoForm, FiltroPagos, FiltroPendientes
from django.contrib.auth.decorators import login_required
from core.autenticacion import grupo_requerido
from control_morosidad.views import reconectarClienteEspecifico
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.http import HttpResponse
from django.conf import settings
import os

# Create your views here.

# Este metodo se encarga de mostrar la lista de pagos registrados, 
# con la posibilidad de aplicar filtros por nombre del cliente y rango de fechas.
@login_required
@grupo_requerido('asistente_administrativo')
def gestion_pago(request,id):

    if id != 0:
        todos_pagos = Pago.objects.filter(idCliente__id=id).order_by('-fecha')
    else:
        todos_pagos = Pago.objects.all().order_by('-fecha')

    if request.method == 'POST':
        filtro = FiltroPagos(request.POST)
        if filtro.is_valid():
            nombreCliente = filtro.cleaned_data.get('nombreCliente')
            fecha_inicio = filtro.cleaned_data.get('fecha_inicio')
            fecha_fin = filtro.cleaned_data.get('fecha_fin')

            if nombreCliente:
                todos_pagos = todos_pagos.filter(Q(idCliente__nombre__icontains=nombreCliente) | Q(idCliente__cedula__istartswith=nombreCliente))

            if fecha_inicio and fecha_fin:
                todos_pagos = todos_pagos.filter(fecha__range=(fecha_inicio, fecha_fin))
            elif fecha_inicio:
                todos_pagos = todos_pagos.filter(fecha__gte=fecha_inicio)
            elif fecha_fin:
                todos_pagos = todos_pagos.filter(fecha__lte=fecha_fin)

            todos_pagos = todos_pagos.order_by('-fecha')
    else:
        filtro = FiltroPagos()

    paginator = Paginator(todos_pagos, 10)
    query_params = request.GET.copy()

    if 'page' in query_params:
        del query_params['page']
    
    pagos = paginator.get_page(request.GET.get('page'))
        
    return render(request, 'gestion_pagos.html', {'pagos': pagos, 'filtros': filtro, 'query_string': query_params.urlencode()})


# Este metodo se encarga de registrar un nuevo pago para un cliente específico, 
# actualizando el saldo del cliente y su estado si es necesario y si las entradas son validas. 
# Registra un log detallado del nuevo pago registrado, incluyendo el monto y la tasa aplicada.
# Reconecta al cliente si su saldo llega a cero después del pago.
@login_required
@grupo_requerido('asistente_administrativo')
def crear_pago(request,id):
    cliente = get_object_or_404(Cliente,borrado=False, id=id)

    if request.method == 'POST':
        form = PagoForm(request.POST, request.FILES)

        if form.is_valid():
            montoUSD = form.cleaned_data.get('montoUSD')
            cliente.saldo -= montoUSD

            if cliente.saldo < 0:
                cliente.saldo += montoUSD
                return render(request, 'crear_pago.html', {
                    'form': form, 
                    'cliente': cliente,
                    'fecha': timezone.now(),
                    'error': 'El monto del pago excede el saldo pendiente del cliente.'
                    })
            elif cliente.saldo == 0 and cliente.estado == 'Desconectado':
                if reconectarClienteEspecifico(cliente):
                    cliente.estado = 'Solvente'

            nuevo_pago = form.save(commit=False)
            nuevo_pago.idPersonal = request.user
            nuevo_pago.idCliente = cliente
            nuevo_pago.save()

            cliente.save()

            Logs.objects.create(
                idPersonal=request.user,
                mensaje=f"""Registró un nuevo pago para el cliente {cliente.nombre} (Cédula: {cliente.cedula}). 
Monto: {montoUSD}$
Tasa: {nuevo_pago.tasa}""",
                modulo = "Gestion de pagos",
                error = False,
                fecha = timezone.now()
            )
            return redirect('gestion_pagos',0)
    else:
        form = PagoForm()
    
    return render(request, 'crear_pago.html', {'form': form, 'cliente': cliente, 'fecha': timezone.now()})


# Este metodo se encarga de modificar un pago existente, 
# actualizando el saldo del cliente y su estado si es necesario y si las entradas son validas, 
# y registrando un log detallado de los cambios realizados.
@login_required
@grupo_requerido('asistente_administrativo')
def modificar_pago(request, id):
    pago = get_object_or_404(Pago, id=id)
    cliente = pago.idCliente
    pago_anterior = Pago.objects.get(id=id)
    montoAnterior = pago.montoUSD

    if request.method == 'POST':
        form = PagoForm(request.POST, request.FILES, instance=pago)
        
        if form.is_valid():
            cliente = pago.idCliente
            montoNuevo = form.cleaned_data.get('montoUSD')            
            diferencia = montoNuevo - montoAnterior
            cliente.saldo -= diferencia
                
            if cliente.saldo < 0:
                cliente.saldo += diferencia
                return render(request, 'modificar_pago.html', {
                    'form': form, 
                    'cliente': cliente, 
                    'fecha': pago.fecha,
                    'error': 'El monto del pago excede el saldo pendiente del cliente.'
                })
            elif cliente.saldo == 0 and cliente.estado == 'Desconectado':
                if reconectarClienteEspecifico(cliente):
                    cliente.estado = 'Solvente'
            else:
                cliente.estado = 'Pendiente'

            cliente.save()

            pago = form.save(commit=False)
            pago.idPersonal = request.user

            lista_cambios = []

            for campo, valor_nuevo in pago.__dict__.items():

                if campo.startswith('_') or campo in ['id', 'borrado']:
                    continue
                
                valor_viejo = getattr(pago_anterior, campo)
                
                if valor_nuevo != valor_viejo:
                    lista_cambios.append(f"{campo}: {valor_viejo} => {valor_nuevo}")

            if lista_cambios:
                mensaje_log = f"Modificó al pago (ID: {pago.id}). Cambios realizados:\n\n" + "\n".join(lista_cambios)
            else:
                mensaje_log = f"El operador guardó al pago (ID: {pago.id}) sin realizar cambios."
            
            pago.save(update_fields=['montoUSD', 'tasa', 'comprobante', 'metodo'])
            form.save_m2m()
            
            Logs.objects.create(
                idPersonal=request.user, 
                mensaje=mensaje_log,
                modulo = "Gestion de pagos",
                error = False,
                fecha = timezone.now()
            )
            
            return redirect('gestion_pagos',0)
    else:
        form = PagoForm(instance=pago)
        
    return render(request, 'modificar_pago.html', {'form': form, 'cliente': cliente, 'fecha': pago.fecha})


# Este metodo se encarga de mostrar los detalles de un pago específico, 
# incluyendo la información del cliente asociado.
@login_required
@grupo_requerido('asistente_administrativo')
def mostrar_detalles(request, id):
    pago = get_object_or_404(Pago, id=id)
    cliente = pago.idCliente
    return render(request, 'detalles_pago.html', {'pago': pago, 'cliente': cliente})


# Este metodo se encarga de mostrar la lista de clientes con pagos pendientes,
# con la posibilidad de aplicar un filtro por nombre del cliente o cedula. 
@login_required
@grupo_requerido('asistente_administrativo')
def pendientes(request):
    clientes_pendientes = Cliente.objects.filter(borrado=False,saldo__gt=0).order_by('nombre')
    filtro = FiltroPendientes()

    if request.method == 'POST':
        filtro = FiltroPendientes(request.POST)
        if filtro.is_valid():
            nombreCliente = filtro.cleaned_data.get('nombreCliente')

            if nombreCliente:
                clientes_pendientes = clientes_pendientes.filter(Q(nombre__icontains=nombreCliente) | Q(cedula__icontains=nombreCliente)).order_by('nombre')

    paginator = Paginator(clientes_pendientes, 10)
    query_params = request.GET.copy()

    if 'page' in query_params:
        del query_params['page']

    clientes = paginator.get_page(request.GET.get('page'))

    return render(request, 'pendientes.html', {'clientes': clientes, 'filtros': filtro, 'query_string': query_params.urlencode()})

# Genera un reporte formateado en Excel con columna de margen vacía, centrado absoluto y el logotipo Logo-1.png real
@login_required
@grupo_requerido('asistente_administrativo')
def exportar_pagos_excel(request, id):
    if id != 0:
        todos_pagos = Pago.objects.filter(idCliente__id=id).order_by('-fecha')
    else:
        todos_pagos = Pago.objects.all().order_by('-fecha')

    if request.method == 'POST':
        filtro = FiltroPagos(request.POST)
        if filtro.is_valid():
            nombreCliente = filtro.cleaned_data.get('nombreCliente')
            fecha_inicio = filtro.cleaned_data.get('fecha_inicio')
            fecha_fin = filtro.cleaned_data.get('fecha_fin')

            if nombreCliente:
                todos_pagos = todos_pagos.filter(
                    Q(idCliente__nombre__icontains=nombreCliente) | 
                    Q(idCliente__cedula__istartswith=nombreCliente)
                )

            if fecha_inicio and fecha_fin:
                todos_pagos = todos_pagos.filter(fecha__range=(fecha_inicio, fecha_fin))
            elif fecha_inicio:
                todos_pagos = todos_pagos.filter(fecha__gte=fecha_inicio)
            elif fecha_fin:
                todos_pagos = todos_pagos.filter(fecha__lte=fecha_fin)

    todos_pagos = todos_pagos.order_by('-fecha')

    # Instanciar el libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Pagos"
    ws.views.sheetView[0].showGridLines = True

    # Definición de Estilos y Alineación Única Centrada
    fuente_titulo = Font(name='Arial', size=14, bold=True, color='003366')
    fuente_cabecera = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    fuente_datos = Font(name='Arial', size=11)
    
    fill_cabecera = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
    alineacion_centro = Alignment(horizontal='center', vertical='center')
    
    borde_delgado = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    # --- INSERTAR EL LOGOTIPO  ---
    dir_actual = os.path.dirname(os.path.abspath(__file__))
    raiz_app = os.path.abspath(os.path.join(dir_actual, '..', '..'))
    ruta_logo = os.path.join(raiz_app, 'static', 'images', 'Logo-1.png')

    # Definimos alturas proporcionales para el espacio del encabezado
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30

    # Combinamos el bloque entero de 2 filas por 4 columnas
    ws.merge_cells('B1:E2')

    if os.path.exists(ruta_logo):
        try:
            img = OpenpyxlImage(ruta_logo)
            
            # Ajustamos un tamaño más generoso para que llene el espacio combinado
            img.width = 320
            img.height = 76
            
            # Al insertarla en B1 (que está combinada con E2), openpyxl la centrará en el bloque
            ws.add_image(img, 'B1')
            print("LOG: ¡Imagen MikroCore centrada con éxito!")
        except Exception as e:
            print(f"LOG: Error al procesar la imagen: {e}")
    else:
        print(f"LOG: No se encontró la imagen en la ruta calculada: {ruta_logo}")

    # Configuración del Título Principal 
    ws['B3'] = "HISTORIAL DE PAGOS ADMINISTRADOS"
    ws['B3'].font = fuente_titulo
    ws['B3'].alignment = alineacion_centro
    ws.merge_cells('B3:E3')
    ws.row_dimensions[3].height = 30

    # Configuración de las Cabeceras de la Tabla (Inician en la columna B)
    cabeceras = ['Fecha', 'RIF / Cédula', 'Cliente', 'Monto USD']
    ws.row_dimensions[5].height = 25

    for col_num, cabecera in enumerate(cabeceras, 2):
        celda = ws.cell(row=5, column=col_num, value=cabecera)
        celda.font = fuente_cabecera
        celda.fill = fill_cabecera
        celda.alignment = alineacion_centro

    # Llenado dinámico de datos estructurados (De la columna B a la E)
    fila_actual = 6
    for pago in todos_pagos:
        fecha_str = pago.fecha.strftime('%d/%m/%Y') if pago.fecha else ""
        
        ws.cell(row=fila_actual, column=2, value=fecha_str)
        ws.cell(row=fila_actual, column=3, value=pago.idCliente.cedula)
        ws.cell(row=fila_actual, column=4, value=pago.idCliente.nombre)
        ws.cell(row=fila_actual, column=5, value=float(pago.montoUSD))
        
        ws.row_dimensions[fila_actual].height = 20
        
        for col_num in range(2, 6):
            c = ws.cell(row=fila_actual, column=col_num)
            c.font = fuente_datos
            c.border = borde_delgado
            c.alignment = alineacion_centro

        ws.cell(row=fila_actual, column=5).number_format = '$#,##0.00'
        fila_actual += 1

    # Forzar que la columna A sea un margen delgado y limpio
    ws.column_dimensions['A'].width = 5

    # Autoajuste controlado únicamente de las columnas de datos
    columnas_tabla = ['B', 'C', 'D', 'E']
    for col_letter in columnas_tabla:
        max_len = 0
        for row in range(5, fila_actual):
            val_celda = str(ws[f'{col_letter}{row}'].value or '')
            if len(val_celda) > max_len:
                max_len = len(val_celda)
        
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    # Respuesta de descarga binaria
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_pagos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response